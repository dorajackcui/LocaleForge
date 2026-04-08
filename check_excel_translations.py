from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

try:
    import requests
except ImportError:  # pragma: no cover - fallback for machines without requests
    requests = None
    import urllib.request


STATUS_OK = "OK"
STATUS_EMPTY = "EMPTY"
STATUS_SUSPECT = "\u7591\u4f3c\u82f1\u6587\u672a\u7ffb\u8bd1"
DEFAULT_INPUT_NAME = "\u670d\u88c5\u76f8\u5173.xlsx"
DEFAULT_HEADER = "CheckResult"
DEFAULT_PROMPT_FILE = "translation_checker_prompt.txt"
PROMPT_STATUS_OK = "{{STATUS_OK}}"
PROMPT_STATUS_SUSPECT = "{{STATUS_SUSPECT}}"
PROMPT_TEXT = "{{TEXT}}"
REQUIRED_PROMPT_MARKERS = (
    PROMPT_STATUS_OK,
    PROMPT_STATUS_SUSPECT,
    PROMPT_TEXT,
)

FRENCH_MARKERS = {
    "alors",
    "au",
    "aucun",
    "aucune",
    "aux",
    "avec",
    "ce",
    "cet",
    "cette",
    "comme",
    "dans",
    "de",
    "des",
    "du",
    "elle",
    "en",
    "entre",
    "est",
    "et",
    "fait",
    "il",
    "ils",
    "je",
    "la",
    "le",
    "les",
    "leur",
    "lui",
    "mais",
    "mon",
    "ne",
    "ou",
    "par",
    "pas",
    "plus",
    "pour",
    "que",
    "qui",
    "sa",
    "sans",
    "se",
    "ses",
    "son",
    "sont",
    "sur",
    "te",
    "toi",
    "tres",
    "une",
    "un",
    "vers",
    "votre",
    "vos",
}

ENGLISH_MARKERS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "but",
    "by",
    "for",
    "from",
    "in",
    "into",
    "is",
    "it",
    "its",
    "of",
    "on",
    "or",
    "that",
    "the",
    "their",
    "there",
    "these",
    "this",
    "to",
    "was",
    "were",
    "with",
}

FRENCH_PREFIXES = ("l'", "d'", "j'", "n'", "c'", "s'", "m'", "t'", "qu'")
TOKEN_RE = re.compile(r"[A-Za-z\u00C0-\u00FF]+(?:'[A-Za-z\u00C0-\u00FF]+)?")
ASCII_WORD_RE = re.compile(r"[A-Za-z]+(?:'[A-Za-z]+)?")
ACCENT_RE = re.compile(
    r"[\u00E0\u00E2\u00E4\u00E6\u00E7\u00E9\u00E8\u00EA\u00EB\u00EE\u00EF\u00F4\u0153\u00F9\u00FB\u00FC\u00FF]",
    re.IGNORECASE,
)


@dataclass
class RuleDecision:
    status: Optional[str]
    reason: str


@dataclass
class ClassificationResult:
    status: str
    spans: list[str]


ProgressCallback = Callable[[int, int, int, Dict[str, int]], None]


class OllamaClient:
    def __init__(self, api_url: str, model: str, timeout: float, prompt_template: str) -> None:
        self.api_url = api_url.rstrip("/")
        self.model = model
        self.timeout = timeout
        self.prompt_template = prompt_template
        self.session = requests.Session() if requests is not None else None

    def ensure_available(self) -> None:
        tags_url = f"{self.api_url}/api/tags"
        try:
            payload = self._get_json(tags_url)
        except Exception as exc:  # pragma: no cover - depends on local service
            raise RuntimeError(
                f"Cannot reach local Ollama service at {tags_url}. "
                "Please make sure `ollama serve` is running."
            ) from exc

        models = {item.get("name") for item in payload.get("models", [])}
        if self.model not in models:
            raise RuntimeError(
                f"Model `{self.model}` was not found in the local Ollama service. "
                "Run `ollama list` to verify the installed models."
            )

    def classify(self, text: str) -> ClassificationResult:
        prompt = render_prompt(self.prompt_template, text)
        payload = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "format": "json",
            "options": {
                "temperature": 0,
            },
        }
        response = self._post_json(f"{self.api_url}/api/generate", payload)
        raw = response.get("response", "")
        result = self._parse_result(raw)
        if result is None:
            raise RuntimeError(f"Model returned an unparseable response: {raw!r}")
        return result

    def _parse_result(self, raw: str) -> Optional[ClassificationResult]:
        raw = raw.strip()
        candidates = [raw]
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if match:
            candidates.append(match.group(0))

        for item in candidates:
            try:
                parsed = json.loads(item)
            except json.JSONDecodeError:
                continue
            status = str(parsed.get("status", "")).strip()
            if status in {STATUS_OK, STATUS_SUSPECT}:
                spans = self._normalize_spans(parsed.get("spans"))
                return ClassificationResult(status=status, spans=spans)

        if STATUS_SUSPECT in raw:
            return ClassificationResult(status=STATUS_SUSPECT, spans=[])
        if STATUS_OK in raw:
            return ClassificationResult(status=STATUS_OK, spans=[])
        return None

    def _normalize_spans(self, spans_value: object) -> list[str]:
        if not isinstance(spans_value, list):
            return []
        normalized: list[str] = []
        for item in spans_value:
            text = normalize_text(item)
            if text and text not in normalized:
                normalized.append(text)
        return normalized

    def _get_json(self, url: str) -> dict:
        if self.session is not None:
            response = self.session.get(url, timeout=self.timeout)
            response.raise_for_status()
            return response.json()

        request = urllib.request.Request(url=url, method="GET")
        with urllib.request.urlopen(request, timeout=self.timeout) as response:
            return json.loads(response.read().decode("utf-8"))

    def _post_json(self, url: str, payload: dict) -> dict:
        if self.session is not None:
            response = self.session.post(url, json=payload, timeout=self.timeout)
            response.raise_for_status()
            return response.json()

        body = json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            url=url,
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=self.timeout) as response:
            return json.loads(response.read().decode("utf-8"))


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def analyze_text(text: str) -> RuleDecision:
    if not text:
        return RuleDecision(STATUS_EMPTY, "empty cell")

    tokens = TOKEN_RE.findall(text)
    if not tokens:
        return RuleDecision(None, "non-empty text without latin tokens")

    lowered_tokens = [token.lower() for token in tokens]
    stripped_tokens = [token.strip("'") for token in lowered_tokens]

    english_hits = sum(1 for token in stripped_tokens if token in ENGLISH_MARKERS)
    french_hits = sum(1 for token in stripped_tokens if token in FRENCH_MARKERS)
    french_prefix_hits = sum(
        1 for token in lowered_tokens if token.startswith(FRENCH_PREFIXES)
    )
    accented_hits = len(ACCENT_RE.findall(text))
    ascii_words = ASCII_WORD_RE.findall(text)
    ascii_ratio = len(ascii_words) / max(len(tokens), 1)
    token_count = len(tokens)

    if english_hits >= 2 and french_hits == 0 and french_prefix_hits == 0:
        return RuleDecision(STATUS_SUSPECT, "multiple english markers")

    if (
        english_hits >= 1
        and token_count >= 5
        and french_hits == 0
        and french_prefix_hits == 0
        and accented_hits == 0
        and ascii_ratio > 0.8
    ):
        return RuleDecision(STATUS_SUSPECT, "english-heavy sentence")

    # Do not auto-approve French-looking rows.
    # Mixed rows can be mostly French while still leaking one untranslated phrase.
    # Everything that is not empty or clearly English-heavy should go to the model.
    if french_hits >= 1 or french_prefix_hits >= 1 or accented_hits >= 1:
        return RuleDecision(None, "french-looking text requires model review")

    return RuleDecision(None, "fallback to model review")


def default_input_path() -> Path:
    preferred = Path.cwd() / DEFAULT_INPUT_NAME
    if preferred.exists():
        return preferred.resolve()

    matches = sorted(Path.cwd().glob("*.xlsx"))
    if matches:
        return matches[0].resolve()
    return preferred.resolve()


def default_prompt_path() -> Path:
    return Path(__file__).with_name(DEFAULT_PROMPT_FILE).resolve()


def load_prompt_template(prompt_path: Path) -> str:
    resolved = Path(prompt_path).expanduser().resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"Prompt file does not exist: {resolved}")

    template = resolved.read_text(encoding="utf-8").strip()
    if not template:
        raise ValueError(f"Prompt file is empty: {resolved}")

    missing = [marker for marker in REQUIRED_PROMPT_MARKERS if marker not in template]
    if missing:
        missing_list = ", ".join(missing)
        raise ValueError(
            f"Prompt file `{resolved}` is missing required placeholders: {missing_list}"
        )
    return template


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_checked{input_path.suffix}")


def render_prompt(template: str, text: str) -> str:
    prompt = template
    replacements = {
        PROMPT_STATUS_OK: STATUS_OK,
        PROMPT_STATUS_SUSPECT: STATUS_SUSPECT,
        PROMPT_TEXT: text,
    }
    for marker, value in replacements.items():
        prompt = prompt.replace(marker, value)
    return prompt


def get_workbook_sheet_names(input_path: Path) -> list[str]:
    workbook = load_workbook(input_path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def process_workbook(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
    source_col: str,
    result_col: str,
    start_row: int,
    client: OllamaClient,
    progress_callback: Optional[ProgressCallback] = None,
) -> Tuple[int, Dict[str, int]]:
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise KeyError(f"Worksheet `{sheet_name}` not found. Available sheets: {available}")

    worksheet = workbook[sheet_name]
    source_index = column_index_from_string(source_col)
    result_index = column_index_from_string(result_col)
    spans_index = result_index + 1

    if not worksheet.cell(row=1, column=result_index).value:
        worksheet.cell(row=1, column=result_index).value = DEFAULT_HEADER
    if not worksheet.cell(row=1, column=spans_index).value:
        worksheet.cell(row=1, column=spans_index).value = f"{DEFAULT_HEADER}Spans"

    cache: Dict[str, ClassificationResult] = {}
    stats = {
        STATUS_OK: 0,
        STATUS_EMPTY: 0,
        STATUS_SUSPECT: 0,
        "MODEL_CALLS": 0,
        "CACHE_HITS": 0,
    }

    total_rows = max(worksheet.max_row - start_row + 1, 0)
    for offset, row_idx in enumerate(range(start_row, worksheet.max_row + 1), start=1):
        raw_value = worksheet.cell(row=row_idx, column=source_index).value
        text = normalize_text(raw_value)

        decision = analyze_text(text)
        result: Optional[ClassificationResult] = None

        if decision.status is not None:
            result = ClassificationResult(status=decision.status, spans=[])
        else:
            cached = cache.get(text)
            if cached is not None:
                result = cached
                stats["CACHE_HITS"] += 1
            else:
                result = client.classify(text)
                cache[text] = result
                stats["MODEL_CALLS"] += 1

        if result is None:
            raise RuntimeError(f"Failed to classify row {row_idx}")

        worksheet.cell(row=row_idx, column=result_index).value = result.status
        worksheet.cell(row=row_idx, column=spans_index).value = " | ".join(result.spans)
        stats[result.status] += 1

        if progress_callback is not None:
            progress_callback(offset, total_rows, row_idx, dict(stats))

    workbook.save(output_path)
    return total_rows, stats


def cli_progress(offset: int, total_rows: int, row_idx: int, stats: Dict[str, int]) -> None:
    if offset % 100 != 0 and offset != total_rows:
        return
    print(
        f"[{offset}/{total_rows}] row={row_idx} "
        f"OK={stats[STATUS_OK]} "
        f"SUSPECT={stats[STATUS_SUSPECT]} "
        f"EMPTY={stats[STATUS_EMPTY]} "
        f"MODEL={stats['MODEL_CALLS']}"
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Check whether a French translation column still contains untranslated English."
    )
    parser.add_argument(
        "--input",
        default=str(default_input_path()),
        help="Path to the source Excel file.",
    )
    parser.add_argument(
        "--output",
        help="Path to the output Excel file. Defaults to <input>_checked.xlsx.",
    )
    parser.add_argument(
        "--sheet",
        default="Sheet1",
        help="Worksheet name to process.",
    )
    parser.add_argument(
        "--source-col",
        default="C",
        help="Column to inspect. Default: C",
    )
    parser.add_argument(
        "--result-col",
        default="F",
        help="Column to write results to. Default: F",
    )
    parser.add_argument(
        "--model",
        default="gemma4:e4b",
        help="Ollama model name. Default: gemma4:e4b",
    )
    parser.add_argument(
        "--prompt-file",
        default=str(default_prompt_path()),
        help="Path to the prompt template file.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="First row to process. Default: 2",
    )
    parser.add_argument(
        "--api-url",
        default="http://127.0.0.1:11434",
        help="Base URL of the Ollama API.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=120.0,
        help="HTTP timeout in seconds for Ollama API calls.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else default_output_path(input_path)
    )
    prompt_path = Path(args.prompt_file).expanduser().resolve()

    if not input_path.exists():
        print(f"Input file does not exist: {input_path}", file=sys.stderr)
        return 1

    prompt_template = load_prompt_template(prompt_path)
    client = OllamaClient(
        api_url=args.api_url,
        model=args.model,
        timeout=args.timeout,
        prompt_template=prompt_template,
    )
    client.ensure_available()

    print(f"Input : {input_path}")
    print(f"Output: {output_path}")
    print(f"Sheet : {args.sheet}")
    print(f"Model : {args.model}")
    print(f"Prompt: {prompt_path}")

    total_rows, stats = process_workbook(
        input_path=input_path,
        output_path=output_path,
        sheet_name=args.sheet,
        source_col=args.source_col,
        result_col=args.result_col,
        start_row=args.start_row,
        client=client,
        progress_callback=cli_progress,
    )

    print("\nFinished.")
    print(f"Rows processed : {total_rows}")
    print(f"{STATUS_OK:<12}: {stats[STATUS_OK]}")
    print(f"{STATUS_SUSPECT:<12}: {stats[STATUS_SUSPECT]}")
    print(f"{STATUS_EMPTY:<12}: {stats[STATUS_EMPTY]}")
    print(f"MODEL_CALLS  : {stats['MODEL_CALLS']}")
    print(f"CACHE_HITS   : {stats['CACHE_HITS']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
