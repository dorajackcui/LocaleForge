from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from localeforge.config.tasks import STATUS_EMPTY, STATUS_OK, STATUS_SUSPECT, STATUS_TERM_EXTRACTED, get_task_config
from localeforge.types import ClassificationResult
from localeforge.workbook import process_workbook


class FakeClient:
    def __init__(self, responses: dict[str, ClassificationResult]) -> None:
        self.responses = responses
        self.call_count = 0

    def classify(self, text: str) -> ClassificationResult:
        self.call_count += 1
        return self.responses[text]


class WorkbookTests(unittest.TestCase):
    def test_term_extraction_processes_empty_ok_and_extracted_rows(self) -> None:
        task_config = get_task_config("term-extraction")
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "input.xlsx"
            output_path = Path(tmpdir) / "output.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = "Source"
            sheet["A2"] = ""
            sheet["A3"] = "Cast Fireball to restore Mana."
            sheet["A4"] = "Cast Fireball to restore Mana."
            sheet["A5"] = "The weather is nice."
            workbook.save(input_path)

            client = FakeClient(
                {
                    "Cast Fireball to restore Mana.": ClassificationResult(
                        status=STATUS_TERM_EXTRACTED,
                        spans=["Fireball", "Mana"],
                    ),
                    "The weather is nice.": ClassificationResult(
                        status=STATUS_OK,
                        spans=[],
                    ),
                }
            )

            total_rows, stats = process_workbook(
                input_path=input_path,
                output_path=output_path,
                sheet_name="Sheet1",
                source_col="A",
                result_col="B",
                start_row=2,
                client=client,
                task_config=task_config,
            )

            self.assertEqual(total_rows, 4)
            self.assertEqual(client.call_count, 2)
            self.assertEqual(stats[STATUS_EMPTY], 1)
            self.assertEqual(stats[STATUS_OK], 1)
            self.assertEqual(stats[STATUS_TERM_EXTRACTED], 2)
            self.assertEqual(stats["MODEL_CALLS"], 2)
            self.assertEqual(stats["CACHE_HITS"], 1)

            checked = load_workbook(output_path)
            try:
                result_sheet = checked["Sheet1"]
                self.assertEqual(result_sheet["B1"].value, "TermExtractResult")
                self.assertEqual(result_sheet["C1"].value, "ExtractedTerms")
                self.assertEqual(result_sheet["B2"].value, STATUS_EMPTY)
                self.assertIsNone(result_sheet["C2"].value)
                self.assertEqual(result_sheet["B3"].value, STATUS_TERM_EXTRACTED)
                self.assertEqual(result_sheet["C3"].value, "Fireball | Mana")
                self.assertEqual(result_sheet["B4"].value, STATUS_TERM_EXTRACTED)
                self.assertEqual(result_sheet["C4"].value, "Fireball | Mana")
                self.assertEqual(result_sheet["B5"].value, STATUS_OK)
                self.assertIsNone(result_sheet["C5"].value)
            finally:
                checked.close()

    def test_english_check_preserves_rule_and_model_flow(self) -> None:
        task_config = get_task_config("english-check")
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "input.xlsx"
            output_path = Path(tmpdir) / "output.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = "Source"
            sheet["A2"] = ""
            sheet["A3"] = "the hero is in the castle"
            sheet["A4"] = "bonjour le monde"
            workbook.save(input_path)

            client = FakeClient(
                {
                    "bonjour le monde": ClassificationResult(
                        status=STATUS_OK,
                        spans=[],
                    )
                }
            )

            total_rows, stats = process_workbook(
                input_path=input_path,
                output_path=output_path,
                sheet_name="Sheet1",
                source_col="A",
                result_col="B",
                start_row=2,
                client=client,
                task_config=task_config,
            )

            self.assertEqual(total_rows, 3)
            self.assertEqual(client.call_count, 1)
            self.assertEqual(stats[STATUS_EMPTY], 1)
            self.assertEqual(stats[STATUS_SUSPECT], 1)
            self.assertEqual(stats[STATUS_OK], 1)
            self.assertEqual(stats["MODEL_CALLS"], 1)


if __name__ == "__main__":
    unittest.main()

