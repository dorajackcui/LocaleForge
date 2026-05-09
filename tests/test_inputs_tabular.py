from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from localeforge.errors import InputOutputError
from localeforge.inputs import discover_work_items
from localeforge.tabular import load_table


class InputsAndTabularTests(unittest.TestCase):
    def test_single_file_default_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            source = Path(tmpdir) / "a.csv"
            source.write_text("source\nhello\n", encoding="utf-8")

            items = discover_work_items(source)

            self.assertEqual(len(items), 1)
            self.assertEqual(items[0].output.name, "a_localeforge.csv")

    def test_single_file_default_output_can_use_task_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            source = Path(tmpdir) / "mt.localeforge.xlsx"
            source.write_text("source\nhello\n", encoding="utf-8")

            items = discover_work_items(source, output_suffix="review")

            self.assertEqual(len(items), 1)
            self.assertEqual(items[0].output.name, "mt.localeforge_review.xlsx")

    def test_folder_requires_output_dir_and_mirrors_supported_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir) / "raw"
            nested = root / "fr"
            nested.mkdir(parents=True)
            (nested / "a.csv").write_text("source\nhello\n", encoding="utf-8")
            (nested / "skip.txt").write_text("nope", encoding="utf-8")

            with self.assertRaises(InputOutputError):
                discover_work_items(root)

            items = discover_work_items(root, Path(tmpdir) / "out")
            self.assertEqual(len(items), 1)
            self.assertEqual(items[0].output, (Path(tmpdir) / "out" / "fr" / "a_localeforge.csv").resolve())

            task_items = discover_work_items(root, Path(tmpdir) / "out", output_suffix="review")
            self.assertEqual(task_items[0].output, (Path(tmpdir) / "out" / "fr" / "a_review.csv").resolve())

    def test_folder_output_dir_cannot_be_inside_input_tree(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir) / "raw"
            root.mkdir()
            (root / "a.csv").write_text("source\nhello\n", encoding="utf-8")

            with self.assertRaisesRegex(InputOutputError, "outside the input directory"):
                discover_work_items(root, root / "out")

    def test_existing_output_requires_explicit_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            source = Path(tmpdir) / "a.csv"
            output = Path(tmpdir) / "out.csv"
            source.write_text("source\nhello\n", encoding="utf-8")
            output.write_text("existing\n", encoding="utf-8")

            with self.assertRaisesRegex(InputOutputError, "already exists"):
                discover_work_items(source, output_path=output)

            items = discover_work_items(source, output_path=output, allow_overwrite=True)
            self.assertEqual(items[0].output, output.resolve())

    def test_csv_header_matching_and_target_creation(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "a.csv"
            path.write_text("Source,Other\nhello,x\n,y\n", encoding="utf-8", newline="")

            table = load_table(path, sheet_name=None)
            source_col = table.resolve_column("source", create=False)
            target_col = table.resolve_column("target", create=True)

            self.assertEqual(table.get_cell(2, source_col), "hello")
            table.set_cell(2, target_col, "HELLO")
            table.save(Path(tmpdir) / "out.csv")

            with (Path(tmpdir) / "out.csv").open(encoding="utf-8", newline="") as handle:
                rows = list(csv.reader(handle))
            self.assertEqual(rows[0], ["Source", "Other", "target"])
            self.assertEqual(rows[1][2], "HELLO")

    def test_xlsx_header_matching_and_target_creation(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "a.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "source"
            ws["A2"] = "hello"
            wb.save(path)

            table = load_table(path, sheet_name="Sheet1")
            source_col = table.resolve_column("source", create=False)
            target_col = table.resolve_column("target", create=True)
            table.set_cell(2, target_col, "HELLO")
            output = Path(tmpdir) / "out.xlsx"
            table.save(output)

            checked = load_workbook(output)
            try:
                self.assertEqual(checked["Sheet1"]["B1"].value, "target")
                self.assertEqual(checked["Sheet1"]["B2"].value, "HELLO")
            finally:
                checked.close()


if __name__ == "__main__":
    unittest.main()
