from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import column_index_from_string

from .config.tasks import DEFAULT_INPUT_NAME, STATUS_EMPTY, STATUS_OK, TaskConfig
from .rules import get_rule_decision, normalize_text
from .types import ClassificationResult, Classifier, ProgressCallback


TERM_SUMMARY_HEADERS = ("ExtractedTerm", "Occurrences", "SourceRows")


@dataclass
class TermSummaryEntry:
    count: int = 0
    rows: list[int] = field(default_factory=list)


def default_input_path() -> Path:
    preferred = Path.cwd() / DEFAULT_INPUT_NAME
    if preferred.exists():
        return preferred.resolve()

    matches = sorted(Path.cwd().glob("*.xlsx"))
    if matches:
        return matches[0].resolve()
    return preferred.resolve()


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_checked{input_path.suffix}")


def get_workbook_sheet_names(input_path: Path) -> list[str]:
    workbook = load_workbook(input_path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _normalize_term(term: str) -> str:
    return " ".join(term.split()).strip()


def _collect_term_summary(
    summary: dict[str, TermSummaryEntry],
    row_idx: int,
    spans: list[str],
) -> None:
    row_terms: list[str] = []
    seen_terms: set[str] = set()
    for span in spans:
        normalized = _normalize_term(span)
        if not normalized or normalized in seen_terms:
            continue
        seen_terms.add(normalized)
        row_terms.append(normalized)

    for term in row_terms:
        entry = summary.setdefault(term, TermSummaryEntry())
        entry.count += 1
        entry.rows.append(row_idx)


def _write_term_summary_sheet(
    workbook: Workbook,
    sheet_name: str,
    summary: dict[str, TermSummaryEntry],
) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(title=sheet_name)
    for column_idx, header in enumerate(TERM_SUMMARY_HEADERS, start=1):
        sheet.cell(row=1, column=column_idx).value = header

    for row_idx, (term, entry) in enumerate(summary.items(), start=2):
        rows = ", ".join(str(row_number) for row_number in entry.rows)
        sheet.cell(row=row_idx, column=1).value = term
        sheet.cell(row=row_idx, column=2).value = entry.count
        sheet.cell(row=row_idx, column=3).value = rows


def process_workbook(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
    source_col: str,
    result_col: str,
    start_row: int,
    client: Classifier,
    task_config: TaskConfig,
    concurrency: int = 1,
    progress_callback: ProgressCallback | None = None,
) -> tuple[int, dict[str, int]]:
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise KeyError(f"Worksheet `{sheet_name}` not found. Available sheets: {available}")

    worksheet = workbook[sheet_name]
    source_index = column_index_from_string(source_col)
    result_index = column_index_from_string(result_col)
    spans_index = result_index + 1

    if not worksheet.cell(row=1, column=result_index).value:
        worksheet.cell(row=1, column=result_index).value = task_config.result_header
    if not worksheet.cell(row=1, column=spans_index).value:
        worksheet.cell(row=1, column=spans_index).value = task_config.details_header

    row_results: dict[int, ClassificationResult] = {}
    pending_rows: dict[str, list[int]] = {}
    term_summary: dict[str, TermSummaryEntry] = {}
    stats = {
        STATUS_OK: 0,
        STATUS_EMPTY: 0,
        task_config.hit_status: 0,
        "MODEL_CALLS": 0,
        "CACHE_HITS": 0,
    }

    total_rows = max(worksheet.max_row - start_row + 1, 0)
    completed_rows = 0
    for row_idx in range(start_row, worksheet.max_row + 1):
        raw_value = worksheet.cell(row=row_idx, column=source_index).value
        text = normalize_text(raw_value)

        decision = get_rule_decision(task_config, text)

        if decision.status is not None:
            result = ClassificationResult(status=decision.status, spans=[])
            row_results[row_idx] = result
            stats[result.status] += 1
            completed_rows += 1
            if progress_callback is not None:
                progress_callback(completed_rows, total_rows, row_idx, dict(stats))
        else:
            if text in pending_rows:
                pending_rows[text].append(row_idx)
                stats["CACHE_HITS"] += 1
            else:
                pending_rows[text] = [row_idx]

    if pending_rows:
        max_workers = max(1, min(concurrency, len(pending_rows)))
        if max_workers == 1:
            resolved_results = {
                text: client.classify(text)
                for text in pending_rows
            }
        else:
            resolved_results: dict[str, ClassificationResult] = {}
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_text = {executor.submit(client.classify, text): text for text in pending_rows}
                for future in as_completed(future_to_text):
                    text = future_to_text[future]
                    resolved_results[text] = future.result()

        for text, result in resolved_results.items():
            stats["MODEL_CALLS"] += 1
            for row_idx in pending_rows[text]:
                row_results[row_idx] = result
                stats[result.status] += 1
                if task_config.summary_sheet_name is not None and result.spans:
                    _collect_term_summary(term_summary, row_idx, result.spans)
                completed_rows += 1
                if progress_callback is not None:
                    progress_callback(completed_rows, total_rows, row_idx, dict(stats))

    for row_idx in range(start_row, worksheet.max_row + 1):
        result = row_results.get(row_idx)
        if result is None:
            raise RuntimeError(f"Failed to classify row {row_idx}")

        worksheet.cell(row=row_idx, column=result_index).value = result.status
        worksheet.cell(row=row_idx, column=spans_index).value = " | ".join(result.spans)

    if task_config.summary_sheet_name is not None:
        _write_term_summary_sheet(workbook, task_config.summary_sheet_name, term_summary)

    workbook.save(output_path)
    return total_rows, stats
