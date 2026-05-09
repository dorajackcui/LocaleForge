from __future__ import annotations

import csv
import re
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .errors import InputOutputError


EXCEL_COLUMN_RE = re.compile(r"^[A-Za-z]{1,3}$")


class Table(ABC):
    header_row: int

    @property
    @abstractmethod
    def max_row(self) -> int:
        raise NotImplementedError

    @abstractmethod
    def header_values(self) -> list[str]:
        raise NotImplementedError

    @abstractmethod
    def get_cell(self, row: int, column: int) -> str:
        raise NotImplementedError

    @abstractmethod
    def set_cell(self, row: int, column: int, value: str) -> None:
        raise NotImplementedError

    @abstractmethod
    def set_header(self, column: int, value: str) -> None:
        raise NotImplementedError

    @abstractmethod
    def save(self, path: Path) -> None:
        raise NotImplementedError

    def resolve_column(self, selector: str | int, create: bool = False) -> int:
        selector_text = str(selector).strip()
        if not selector_text:
            raise InputOutputError("Column selector cannot be empty.")

        headers = self.header_values()
        normalized = selector_text.casefold()
        for index, header in enumerate(headers, start=1):
            if header.casefold() == normalized:
                return index

        if selector_text.isdigit():
            column = int(selector_text)
            if column < 1:
                raise InputOutputError(f"Column index must be positive: {selector_text}")
            if create:
                self.set_header(column, selector_text)
                return column
            self._ensure_existing_index(column, selector_text)
            return column

        if EXCEL_COLUMN_RE.match(selector_text):
            column = column_index_from_string(selector_text.upper())
            if create:
                self.set_header(column, selector_text)
                return column
            self._ensure_existing_index(column, selector_text)
            return column

        if create:
            column = len(headers) + 1
            self.set_header(column, selector_text)
            return column

        available = ", ".join(header for header in headers if header) or "<none>"
        raise InputOutputError(f"Column `{selector_text}` was not found. Available headers: {available}")

    def _ensure_existing_index(self, column: int, selector: str) -> None:
        if column > len(self.header_values()):
            available = ", ".join(self.header_values()) or "<none>"
            raise InputOutputError(f"Column `{selector}` was not found. Available headers: {available}")


class CsvTable(Table):
    def __init__(self, path: Path, header_row: int = 1) -> None:
        self.path = path
        self.header_row = header_row
        with path.open(encoding="utf-8-sig", newline="") as handle:
            self.rows = [list(row) for row in csv.reader(handle)]
        while len(self.rows) < header_row:
            self.rows.append([])

    @property
    def max_row(self) -> int:
        return len(self.rows)

    def header_values(self) -> list[str]:
        return [str(value or "").strip() for value in self.rows[self.header_row - 1]]

    def get_cell(self, row: int, column: int) -> str:
        if row < 1 or row > len(self.rows):
            return ""
        values = self.rows[row - 1]
        if column < 1 or column > len(values):
            return ""
        return str(values[column - 1] or "").strip()

    def set_cell(self, row: int, column: int, value: str) -> None:
        self._ensure_cell(row, column)
        self.rows[row - 1][column - 1] = value

    def set_header(self, column: int, value: str) -> None:
        self.set_cell(self.header_row, column, value)

    def save(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, lineterminator="\n")
            writer.writerows(self.rows)

    def _ensure_cell(self, row: int, column: int) -> None:
        while len(self.rows) < row:
            self.rows.append([])
        values = self.rows[row - 1]
        while len(values) < column:
            values.append("")


class XlsxTable(Table):
    def __init__(self, path: Path, sheet_name: str | None = None, header_row: int = 1) -> None:
        self.path = path
        self.header_row = header_row
        self.workbook: Workbook = load_workbook(path)
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                available = ", ".join(self.workbook.sheetnames)
                raise InputOutputError(f"Worksheet `{sheet_name}` not found. Available sheets: {available}")
            self.worksheet: Worksheet = self.workbook[sheet_name]
        else:
            self.worksheet = self.workbook.active

    @property
    def max_row(self) -> int:
        return self.worksheet.max_row

    def header_values(self) -> list[str]:
        return [
            str(self.worksheet.cell(row=self.header_row, column=index).value or "").strip()
            for index in range(1, self.worksheet.max_column + 1)
        ]

    def get_cell(self, row: int, column: int) -> str:
        value: Any = self.worksheet.cell(row=row, column=column).value
        return str(value or "").strip()

    def set_cell(self, row: int, column: int, value: str) -> None:
        self.worksheet.cell(row=row, column=column).value = value

    def set_header(self, column: int, value: str) -> None:
        self.set_cell(self.header_row, column, value)

    def save(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(path)
        self.workbook.close()


def load_table(path: Path | str, sheet_name: str | None = None, header_row: int = 1) -> Table:
    resolved = Path(path)
    suffix = resolved.suffix.lower()
    if suffix == ".csv":
        return CsvTable(resolved, header_row=header_row)
    if suffix == ".xlsx":
        return XlsxTable(resolved, sheet_name=sheet_name, header_row=header_row)
    raise InputOutputError(f"Unsupported input file type `{resolved.suffix}`.")
